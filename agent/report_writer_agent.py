# agent/report_writer_agent.py
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import json

def fill_template_excel(template_path: str, sample_path: str, output_path: str):
    """템플릿에 샘플 데이터를 채워서 최종 보고서를 생성합니다."""
    wb = load_workbook(template_path)
    ws = wb.active

    with open(sample_path, "r", encoding="utf-8") as f:
        sample = json.load(f)

    # 헤더 행에서 컬럼명을 찾고, 해당 컬럼에 샘플 데이터 채우기
    for col_idx in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col_idx).value
        if header in sample:
            ws.cell(row=2, column=col_idx).value = sample[header]

    wb.save(output_path)
    return f"최종 보고서가 {output_path}에 생성되었습니다."

# AutoGen 에이전트용 함수 정의
def report_writer_function(template_path: str, sample_path: str, output_path: str):
    """보고서 작성 에이전트의 메인 함수"""
    try:
        result = fill_template_excel(template_path, sample_path, output_path)
        return {"status": "success", "message": result}
    except Exception as e:
        return {"status": "error", "message": str(e)}

# 시스템 메시지
REPORT_WRITER_SYSTEM_MESSAGE = """당신은 문서 작성 전문가입니다.
템플릿과 샘플 데이터를 기반으로 최종 보고서를 생성하는 역할을 담당합니다.
- 템플릿 구조를 정확히 이해합니다
- 샘플 데이터를 적절한 위치에 배치합니다
- 공공기관 문서의 형식을 준수합니다
- 완성도 높은 최종 보고서를 생성합니다"""
